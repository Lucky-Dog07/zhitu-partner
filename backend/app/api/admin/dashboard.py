"""
管理后台 - 仪表盘API
"""
from fastapi import APIRouter, Depends, Query, Response
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, desc, extract
from typing import List, Dict, Optional
from datetime import datetime, timedelta, date
from pydantic import BaseModel
import io
import csv

from ...core.database import get_db
from ...api.deps import get_current_admin
from ...models.user import User
from ...models.learning_path import LearningPath
from ...models.note import Note
from ...models.interview_question import InterviewQuestion
from ...models.login_log import LoginLog
from ...models.operation_log import OperationLog


router = APIRouter()


# Schemas
class OverviewStats(BaseModel):
    total_users: int
    active_users_today: int
    new_users_this_week: int
    total_learning_paths: int
    total_notes: int
    total_interview_questions: int
    online_users: int  # 最近5分钟活跃


class ActivityItem(BaseModel):
    type: str  # "register", "create_path", "create_note", "login"
    username: str
    description: str
    timestamp: datetime


class SystemStatus(BaseModel):
    database_status: str  # "healthy", "slow", "error"
    api_status: str
    cpu_usage: float
    memory_usage: float
    disk_usage: float
    last_check_time: datetime
    

class QuickStats(BaseModel):
    """快速统计（本周对比上周）"""
    users_this_week: int
    users_last_week: int
    users_growth_rate: float
    paths_this_week: int
    paths_last_week: int
    paths_growth_rate: float
    logins_this_week: int
    logins_last_week: int
    logins_growth_rate: float


class PopularPath(BaseModel):
    id: int
    user_id: int
    username: str
    created_at: datetime


class UserGrowthItem(BaseModel):
    date: str
    new_users: int
    total_users: int


class FeatureUsageData(BaseModel):
    learning_paths: int
    notes: int
    interviews: int
    login_success: int
    login_failed: int


class ActivityHeatmapItem(BaseModel):
    day: str
    hour: int
    count: int


@router.get("/dashboard/overview", response_model=OverviewStats)
async def get_overview(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
):
    """获取仪表盘概览数据"""
    
    # 总用户数
    total_users = db.query(func.count(User.id)).scalar()
    
    # 今日活跃用户（今日登录过）
    today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    active_users_today = db.query(func.count(func.distinct(LoginLog.user_id)))\
        .filter(
            and_(
                LoginLog.login_time >= today_start,
                LoginLog.status == 'success'
            )
        ).scalar()
    
    # 本周新增用户
    week_start = datetime.now() - timedelta(days=7)
    new_users_this_week = db.query(func.count(User.id))\
        .filter(User.created_at >= week_start)\
        .scalar()
    
    # 总学习路线数
    total_learning_paths = db.query(func.count(LearningPath.id)).scalar()
    
    # 总笔记数
    total_notes = db.query(func.count(Note.id)).scalar()
    
    # 总面试题数
    total_interview_questions = db.query(func.count(InterviewQuestion.id)).scalar()
    
    # 在线用户（最近5分钟有登录记录）
    five_minutes_ago = datetime.now() - timedelta(minutes=5)
    online_users = db.query(func.count(func.distinct(LoginLog.user_id)))\
        .filter(
            and_(
                LoginLog.login_time >= five_minutes_ago,
                LoginLog.status == 'success'
            )
        ).scalar()
    
    return OverviewStats(
        total_users=total_users or 0,
        active_users_today=active_users_today or 0,
        new_users_this_week=new_users_this_week or 0,
        total_learning_paths=total_learning_paths or 0,
        total_notes=total_notes or 0,
        total_interview_questions=total_interview_questions or 0,
        online_users=online_users or 0
    )


@router.get("/dashboard/recent-activities")
async def get_recent_activities(
    limit: int = 20,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
) -> List[ActivityItem]:
    """获取最近活动"""
    
    activities = []
    
    # 获取最近注册的用户
    recent_users = db.query(User)\
        .order_by(desc(User.created_at))\
        .limit(5)\
        .all()
    
    for user in recent_users:
        activities.append(ActivityItem(
            type="register",
            username=user.username,
            description=f"新用户 {user.username} 注册了账号",
            timestamp=user.created_at
        ))
    
    # 获取最近创建的学习路线
    recent_paths = db.query(LearningPath)\
        .join(User, LearningPath.user_id == User.id)\
        .order_by(desc(LearningPath.created_at))\
        .limit(5)\
        .all()
    
    for path in recent_paths:
        activities.append(ActivityItem(
            type="create_path",
            username=path.user.username,
            description=f"{path.user.username} 创建了学习路线",
            timestamp=path.created_at
        ))
    
    # 获取最近登录记录
    recent_logins = db.query(LoginLog)\
        .filter(LoginLog.status == 'success')\
        .order_by(desc(LoginLog.login_time))\
        .limit(5)\
        .all()
    
    for login in recent_logins:
        activities.append(ActivityItem(
            type="login",
            username=login.username,
            description=f"{login.username} 登录了系统",
            timestamp=login.login_time
        ))
    
    # 按时间排序
    activities.sort(key=lambda x: x.timestamp, reverse=True)
    
    return activities[:limit]


@router.get("/dashboard/system-status", response_model=SystemStatus)
async def get_system_status(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
):
    """获取系统状态"""
    
    try:
        # 测试数据库连接
        db.execute("SELECT 1")
        database_status = "healthy"
    except Exception:
        database_status = "error"
    
    # 获取系统资源使用情况
    try:
        import psutil
        cpu_usage = psutil.cpu_percent(interval=1)
        memory_info = psutil.virtual_memory()
        disk_info = psutil.disk_usage('/')
    except Exception:
        cpu_usage = 0
        memory_info = type('obj', (object,), {'percent': 0})()
        disk_info = type('obj', (object,), {'percent': 0})()
    
    return SystemStatus(
        database_status=database_status,
        api_status="healthy",
        cpu_usage=cpu_usage,
        memory_usage=memory_info.percent,
        disk_usage=disk_info.percent,
        last_check_time=datetime.now()
    )


@router.get("/dashboard/quick-stats", response_model=QuickStats)
async def get_quick_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
):
    """获取快速统计（本周对比上周）"""
    
    now = datetime.now()
    this_week_start = now - timedelta(days=7)
    last_week_start = now - timedelta(days=14)
    
    # 本周新增用户
    users_this_week = db.query(func.count(User.id))\
        .filter(User.created_at >= this_week_start)\
        .scalar() or 0
    
    # 上周新增用户
    users_last_week = db.query(func.count(User.id))\
        .filter(
            and_(
                User.created_at >= last_week_start,
                User.created_at < this_week_start
            )
        ).scalar() or 0
    
    # 计算增长率
    users_growth_rate = ((users_this_week - users_last_week) / users_last_week * 100) if users_last_week > 0 else 0
    
    # 本周新增学习路线
    paths_this_week = db.query(func.count(LearningPath.id))\
        .filter(LearningPath.created_at >= this_week_start)\
        .scalar() or 0
    
    # 上周新增学习路线
    paths_last_week = db.query(func.count(LearningPath.id))\
        .filter(
            and_(
                LearningPath.created_at >= last_week_start,
                LearningPath.created_at < this_week_start
            )
        ).scalar() or 0
    
    paths_growth_rate = ((paths_this_week - paths_last_week) / paths_last_week * 100) if paths_last_week > 0 else 0
    
    # 本周登录次数
    logins_this_week = db.query(func.count(LoginLog.id))\
        .filter(
            and_(
                LoginLog.login_time >= this_week_start,
                LoginLog.status == 'success'
            )
        ).scalar() or 0
    
    # 上周登录次数
    logins_last_week = db.query(func.count(LoginLog.id))\
        .filter(
            and_(
                LoginLog.login_time >= last_week_start,
                LoginLog.login_time < this_week_start,
                LoginLog.status == 'success'
            )
        ).scalar() or 0
    
    logins_growth_rate = ((logins_this_week - logins_last_week) / logins_last_week * 100) if logins_last_week > 0 else 0
    
    return QuickStats(
        users_this_week=users_this_week,
        users_last_week=users_last_week,
        users_growth_rate=round(users_growth_rate, 2),
        paths_this_week=paths_this_week,
        paths_last_week=paths_last_week,
        paths_growth_rate=round(paths_growth_rate, 2),
        logins_this_week=logins_this_week,
        logins_last_week=logins_last_week,
        logins_growth_rate=round(logins_growth_rate, 2)
    )


@router.get("/dashboard/popular-paths")
async def get_popular_paths(
    limit: int = 5,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
) -> List[PopularPath]:
    """获取热门学习路线"""
    
    # 获取最近创建的路线（作为热门路线的简单实现）
    popular_paths = db.query(LearningPath)\
        .join(User, LearningPath.user_id == User.id)\
        .order_by(desc(LearningPath.created_at))\
        .limit(limit)\
        .all()
    
    return [
        PopularPath(
            id=path.id,
            user_id=path.user_id,
            username=path.user.username,
            created_at=path.created_at
        )
        for path in popular_paths
    ]


@router.get("/dashboard/user-growth", response_model=List[UserGrowthItem])
async def get_user_growth(
    days: int = Query(30, ge=1, le=90),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
):
    """获取用户增长趋势"""
    
    result = []
    end_date = date.today()
    start_date = end_date - timedelta(days=days)
    
    # 获取总用户数的起始值
    total_users_start = db.query(func.count(User.id))\
        .filter(User.created_at < start_date)\
        .scalar() or 0
    
    # 按日期统计新增用户
    for i in range(days):
        current_date = start_date + timedelta(days=i)
        next_date = current_date + timedelta(days=1)
        
        new_users = db.query(func.count(User.id))\
            .filter(
                and_(
                    User.created_at >= current_date,
                    User.created_at < next_date
                )
            ).scalar() or 0
        
        total_users_start += new_users
        
        result.append(UserGrowthItem(
            date=current_date.strftime('%Y-%m-%d'),
            new_users=new_users,
            total_users=total_users_start
        ))
    
    return result


@router.get("/dashboard/feature-usage", response_model=FeatureUsageData)
async def get_feature_usage(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
):
    """获取功能使用分布"""
    
    # 统计各功能使用数量
    learning_paths = db.query(func.count(LearningPath.id)).scalar() or 0
    notes = db.query(func.count(Note.id)).scalar() or 0
    interviews = db.query(func.count(InterviewQuestion.id)).scalar() or 0
    
    # 统计登录成功和失败次数
    login_success = db.query(func.count(LoginLog.id))\
        .filter(LoginLog.status == 'success')\
        .scalar() or 0
    
    login_failed = db.query(func.count(LoginLog.id))\
        .filter(LoginLog.status == 'failed')\
        .scalar() or 0
    
    return FeatureUsageData(
        learning_paths=learning_paths,
        notes=notes,
        interviews=interviews,
        login_success=login_success,
        login_failed=login_failed
    )


@router.get("/dashboard/activity-heatmap", response_model=List[ActivityHeatmapItem])
async def get_activity_heatmap(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
):
    """获取用户活跃度热力图（7天x24小时）"""
    
    result = []
    end_date = datetime.now()
    start_date = end_date - timedelta(days=7)
    
    # 按星期和小时统计登录次数
    days_map = {
        0: 'Monday',
        1: 'Tuesday',
        2: 'Wednesday',
        3: 'Thursday',
        4: 'Friday',
        5: 'Saturday',
        6: 'Sunday'
    }
    
    # 获取所有登录记录
    logins = db.query(
        extract('dow', LoginLog.login_time).label('dow'),
        extract('hour', LoginLog.login_time).label('hour'),
        func.count(LoginLog.id).label('count')
    ).filter(
        and_(
            LoginLog.login_time >= start_date,
            LoginLog.status == 'success'
        )
    ).group_by('dow', 'hour').all()
    
    # 构建热力图数据
    for dow, hour, count in logins:
        # PostgreSQL的dow: 0=Sunday, 1=Monday, ..., 6=Saturday
        # 转换为Python: 0=Monday, ..., 6=Sunday
        day_index = (int(dow) + 6) % 7
        day_name = days_map[day_index]
        
        result.append(ActivityHeatmapItem(
            day=day_name,
            hour=int(hour),
            count=count
        ))
    
    return result


@router.get("/dashboard/export")
async def export_dashboard_data(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    format: str = Query('csv', regex='^(csv|xlsx)$'),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin)
):
    """导出仪表盘数据"""
    
    # 解析日期范围
    if start_date:
        start = datetime.fromisoformat(start_date)
    else:
        start = datetime.now() - timedelta(days=30)
    
    if end_date:
        end = datetime.fromisoformat(end_date)
    else:
        end = datetime.now()
    
    # 收集数据
    data = {
        'users': db.query(User).filter(
            and_(User.created_at >= start, User.created_at <= end)
        ).count(),
        'learning_paths': db.query(LearningPath).filter(
            and_(LearningPath.created_at >= start, LearningPath.created_at <= end)
        ).count(),
        'notes': db.query(Note).filter(
            and_(Note.created_at >= start, Note.created_at <= end)
        ).count(),
        'logins': db.query(LoginLog).filter(
            and_(LoginLog.login_time >= start, LoginLog.login_time <= end)
        ).count()
    }
    
    # 导出为CSV
    if format == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 写入表头
        writer.writerow(['指标', '数值'])
        writer.writerow(['日期范围', f'{start.date()} 至 {end.date()}'])
        writer.writerow(['新增用户', data['users']])
        writer.writerow(['新增学习路线', data['learning_paths']])
        writer.writerow(['新增笔记', data['notes']])
        writer.writerow(['登录次数', data['logins']])
        
        output.seek(0)
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=dashboard_export_{datetime.now().strftime('%Y%m%d')}.csv"
            }
        )
    
    # 导出为Excel (需要openpyxl)
    else:
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "仪表盘数据"
            
            # 写入数据
            ws['A1'] = '指标'
            ws['B1'] = '数值'
            ws['A2'] = '日期范围'
            ws['B2'] = f'{start.date()} 至 {end.date()}'
            ws['A3'] = '新增用户'
            ws['B3'] = data['users']
            ws['A4'] = '新增学习路线'
            ws['B4'] = data['learning_paths']
            ws['A5'] = '新增笔记'
            ws['B5'] = data['notes']
            ws['A6'] = '登录次数'
            ws['B6'] = data['logins']
            
            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=dashboard_export_{datetime.now().strftime('%Y%m%d')}.xlsx"
                }
            )
        except ImportError:
            return {"error": "openpyxl not installed. Please install it to export Excel files."}
